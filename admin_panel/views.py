import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model, authenticate, login
from django.db import transaction
from django.db.models import Sum, Q, Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage
from django.core.mail import send_mail
from django.conf import settings as django_settings
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
import json
import csv
from io import StringIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .decorators import admin_required
from transactions.models import Deposit, Withdrawal, Investment, Transaction, InvestmentPlan
from accounts.models import KYCDocument, UserProfile, CustomUser
from core.models import SiteSettings, Testimonial, Certification
from django.contrib.auth.decorators import login_required, user_passes_test

logger = logging.getLogger(__name__)

User = get_user_model()

def is_admin(user):
    return user.is_authenticated and user.is_staff

def get_or_create_profile(user):
    """Ensure a UserProfile exists for the given user."""
    profile, created = UserProfile.objects.get_or_create(user=user)
    return profile

def admin_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin_panel:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('admin_panel:dashboard')
        else:
            messages.error(request, 'Invalid credentials or insufficient permissions.')
    
    return render(request, 'admin_panel/login.html')

@admin_required
def dashboard(request):
    total_users = User.objects.count()
    pending_deposits = Deposit.objects.filter(status='pending').count()
    pending_withdrawals = Withdrawal.objects.filter(status='pending').count()
    pending_kyc = KYCDocument.objects.filter(status='pending').count()
    
    total_deposits = Deposit.objects.filter(status='approved').aggregate(Sum('amount'))['amount__sum'] or 0
    active_investments = Investment.objects.filter(status='active').aggregate(Sum('amount'))['amount__sum'] or 0
    
    recent_deposits = Deposit.objects.filter(status='pending').order_by('-created_at')[:5]
    recent_withdrawals = Withdrawal.objects.filter(status='pending').order_by('-created_at')[:5]
    
    context = {
        'total_users': total_users,
        'pending_deposits': pending_deposits,
        'pending_withdrawals': pending_withdrawals,
        'pending_kyc': pending_kyc,
        'total_deposits': total_deposits,
        'active_investments': active_investments,
        'recent_deposits': recent_deposits,
        'recent_withdrawals': recent_withdrawals,
    }
    
    return render(request, 'admin_panel/dashboard.html', context)

@admin_required
def manage_users(request):
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')
    date_filter = request.GET.get('date')
    
    users = User.objects.select_related('profile').filter(is_staff=False)
    
    if status_filter == 'pending':
        users = users.filter(
            Q(kyc_documents__status='pending') | 
            Q(kyc_documents__isnull=True)
        ).distinct()
    elif status_filter == 'verified':
        users = users.filter(kyc_documents__status='approved').distinct()
    elif status_filter == 'approved':
        users = users.filter(is_approved=True)
    
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(full_name__icontains=search_query)
        )
    
    if date_filter:
        users = users.filter(date_joined__date=date_filter)
    
    users = users.order_by('-date_joined')
    
    paginator = Paginator(users, 25)
    page_number = request.GET.get('page')
    try:
        users = paginator.get_page(page_number)
    except EmptyPage:
        users = paginator.get_page(1)
    
    context = {
        'users': users,
        'status_filter': status_filter,
        'search_query': search_query,
        'date_filter': date_filter,
    }
    return render(request, 'admin_panel/users.html', context)

@admin_required
def get_user_details(request, user_id):
    try:
        user = get_object_or_404(User, id=user_id)
        profile = get_or_create_profile(user)
        transactions = Transaction.objects.filter(user=user).order_by('-created_at')[:10]
        
        user_data = {
            'id': user.id,
            'username': user.username,
            'full_name': user.full_name,
            'email': user.email,
            'profile': {
                'country': profile.country,
                'account_balance': float(profile.account_balance),
            }
        }
        
        transactions_data = [{
            'transaction_type': t.transaction_type,
            'amount': float(t.amount),
            'created_at': t.created_at.isoformat(),
            'description': t.description
        } for t in transactions]
        
        return JsonResponse({
            'status': 'success',
            'user': user_data,
            'transactions': transactions_data
        })
    except Exception as e:
        logger.error(f'Error fetching user details for user {user_id}: {str(e)}')
        return JsonResponse({
            'status': 'error',
            'message': 'An error occurred while fetching user details.'
        }, status=400)

@admin_required
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        try:
            user.full_name = request.POST.get('full_name')
            user.email = request.POST.get('email')
            profile = get_or_create_profile(user)
            profile.country = request.POST.get('country')
            user.save()
            profile.save()
            messages.success(request, f'User {user.username} updated successfully.')
            return JsonResponse({'status': 'success'})
        except Exception as e:
            logger.error(f'Error editing user {user_id}: {str(e)}')
            return JsonResponse({'status': 'error', 'message': 'An error occurred while updating the user.'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@admin_required
def edit_balance(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                new_balance = float(request.POST.get('account_balance'))
                adjustment_note = request.POST.get('adjustment_note')
                profile = get_or_create_profile(user)
                old_balance = profile.account_balance
                
                profile.account_balance = new_balance
                profile.save()
                
                Transaction.objects.create(
                    user=user,
                    transaction_type='bonus' if new_balance > old_balance else 'withdrawal',
                    amount=abs(new_balance - old_balance),
                    description=f'Admin balance adjustment: {adjustment_note}',
                    reference_id=f'admin_adjust_{user.id}'
                )
                
                messages.success(request, f'Balance for {user.username} updated successfully.')
                return JsonResponse({'status': 'success'})
        except ValueError as e:
            logger.error(f'ValueError editing balance for user {user_id}: {str(e)}')
            return JsonResponse({'status': 'error', 'message': 'Invalid balance value.'}, status=400)
        except Exception as e:
            logger.error(f'Error editing balance for user {user_id}: {str(e)}')
            return JsonResponse({'status': 'error', 'message': 'An error occurred while updating the balance.'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@admin_required
def send_message(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        
        if not subject or not message:
            return JsonResponse({'status': 'error', 'message': 'Subject and message are required'}, status=400)
        
        try:
            send_mail(
                subject,
                message,
                django_settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            messages.success(request, f'Message sent to {user.username} successfully.')
            return JsonResponse({'status': 'success'})
        except Exception as e:
            logger.error(f'Error sending message to user {user_id}: {str(e)}')
            messages.error(request, f'Failed to send message: {str(e)}')
            return JsonResponse({'status': 'error', 'message': 'Failed to send message.'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@admin_required
def view_transactions(request, user_id):
    user = get_object_or_404(User, id=user_id)
    transactions = Transaction.objects.filter(user=user).order_by('-created_at')
    
    context = {
        'user': user,
        'transactions': transactions,
    }
    return render(request, 'admin_panel/user_transactions.html', context)

@admin_required
def login_as_user(request, user_id):
    if not request.user.is_staff:
        messages.error(request, 'Permission denied.')
        return redirect('admin_panel:dashboard')
    
    request.session['admin_user_id'] = request.user.id
    request.session.modified = True
    
    user = get_object_or_404(User, id=user_id)
    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    messages.info(request, f'Logged in as {user.username}. You can switch back to admin from the dashboard.')
    return redirect('core:dashboard')

@admin_required
def switch_back_admin(request):
    admin_user_id = request.session.get('admin_user_id')
    
    if not admin_user_id:
        messages.error(request, 'No admin session found.')
        return redirect('core:dashboard')
    
    try:
        admin_user = get_object_or_404(User, id=admin_user_id)
        if not admin_user.is_staff:
            messages.error(request, 'Invalid admin account.')
            return redirect('core:dashboard')
        
        login(request, admin_user, backend='django.contrib.auth.backends.ModelBackend')
        del request.session['admin_user_id']
        request.session.modified = True
        messages.success(request, f'Switched back to admin account: {admin_user.username}.')
        return redirect('admin_panel:dashboard')
    except Exception as e:
        logger.error(f'Error switching back to admin: {str(e)}')
        messages.error(request, 'An error occurred while switching back to admin.')
        return redirect('core:dashboard')

@admin_required
def suspend_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.is_active = False
        user.save()
        messages.success(request, f'User {user.username} suspended successfully.')
        return redirect('admin_panel:manage_users')
    
    return render(request, 'admin_panel/suspend_user.html', {'user': user})

@admin_required
def approve_kyc(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        try:
            kyc_doc = KYCDocument.objects.filter(user=user).first()
            if kyc_doc:
                kyc_doc.status = 'approved'
                kyc_doc.save()
            
            user.is_verified = True
            user.save()
            
            messages.success(request, f'KYC approved for {user.username}')
        except Exception as e:
            logger.error(f'Error approving KYC for user {user_id}: {str(e)}')
            messages.error(request, f'Error approving KYC: {str(e)}')
    
    return redirect('admin_panel:manage_users')

@admin_required
def reject_kyc(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        try:
            kyc_doc = KYCDocument.objects.filter(user=user).first()
            if kyc_doc:
                kyc_doc.status = 'rejected'
                kyc_doc.save()
            
            messages.success(request, f'KYC rejected for {user.username}')
        except Exception as e:
            logger.error(f'Error rejecting KYC for user {user_id}: {str(e)}')
            messages.error(request, f'Error rejecting KYC: {str(e)}')
    
    return redirect('admin_panel:manage_users')

@admin_required
def manage_deposits(request):
    status_filter = request.GET.get('status')
    deposits = Deposit.objects.select_related('user').all()
    
    if status_filter:
        deposits = deposits.filter(status=status_filter)
    
    deposits = deposits.order_by('-created_at')
    
    context = {
        'deposits': deposits,
        'status_filter': status_filter,
    }
    return render(request, 'admin_panel/deposits.html', context)

@admin_required
def approve_deposit(request, deposit_id):
    deposit = get_object_or_404(Deposit, id=deposit_id)
    
    if deposit.status == 'pending':
        with transaction.atomic():
            deposit.status = 'approved'
            deposit.approved_at = timezone.now()
            deposit.save()
            
            profile = get_or_create_profile(deposit.user)
            profile.account_balance += deposit.amount
            profile.total_deposit += deposit.amount
            profile.save()
            
            Transaction.objects.create(
                user=deposit.user,
                transaction_type='deposit',
                amount=deposit.amount,
                description=f'Deposit approved - {deposit.payment_method}',
                reference_id=str(deposit.id)
            )
            
            messages.success(request, f'Deposit of ${deposit.amount} approved for {deposit.user.username}.')
    
    return redirect('admin_panel:manage_deposits')

@admin_required
def reject_deposit(request, deposit_id):
    deposit = get_object_or_404(Deposit, id=deposit_id)
    
    if deposit.status == 'pending':
        deposit.status = 'rejected'
        deposit.save()
        messages.warning(request, f'Deposit of ${deposit.amount} rejected for {deposit.user.username}.')
    
    return redirect('admin_panel:manage_deposits')

@admin_required
def manage_withdrawals(request):
    status_filter = request.GET.get('status')
    withdrawals = Withdrawal.objects.select_related('user').all()
    
    if status_filter:
        withdrawals = withdrawals.filter(status=status_filter)
    
    withdrawals = withdrawals.order_by('-created_at')
    
    context = {
        'withdrawals': withdrawals,
        'status_filter': status_filter,
    }
    return render(request, 'admin_panel/withdrawals.html', context)

@admin_required
def approve_withdrawal(request, withdrawal_id):
    withdrawal = get_object_or_404(Withdrawal, id=withdrawal_id)
    
    if withdrawal.status == 'pending':
        with transaction.atomic():
            withdrawal.status = 'approved'
            withdrawal.processed_at = timezone.now()
            withdrawal.save()
            
            profile = get_or_create_profile(withdrawal.user)
            profile.account_balance -= withdrawal.amount
            profile.pending_withdrawal -= withdrawal.amount
            profile.total_withdrawal += withdrawal.amount
            profile.save()
            
            messages.success(request, f'Withdrawal of ${withdrawal.amount} approved for {withdrawal.user.username}.')
    
    return redirect('admin_panel:manage_withdrawals')

@login_required
@user_passes_test(is_admin)
def manage_investments(request):
    """Display all investments with filtering and pagination."""
    investments = Investment.objects.select_related('user', 'plan').all()
    
    status_filter = request.GET.get('status')
    plan_filter = request.GET.get('plan')
    search_query = request.GET.get('search')
    export_format = request.GET.get('export')
    
    if status_filter:
        investments = investments.filter(status=status_filter)
    
    if plan_filter:
        investments = investments.filter(plan_id=plan_filter)
    
    if search_query:
        investments = investments.filter(
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__full_name__icontains=search_query)
        )
    
    if export_format in ['csv', 'excel']:
        return export_investments(investments, export_format)
    
    investments = investments.order_by('-start_date')
    
    stats = {
        'total_investment_amount': investments.aggregate(total=Sum('amount'))['total'] or 0,
        'active_investments_count': investments.filter(status='active').count(),
        'completed_investments_count': investments.filter(status='completed').count(),
        'total_expected_returns': investments.aggregate(total=Sum('expected_return'))['total'] or 0,
    }
    
    paginator = Paginator(investments, 25)
    page_number = request.GET.get('page')
    try:
        investments_page = paginator.get_page(page_number)
    except EmptyPage:
        investments_page = paginator.get_page(1)
    
    for investment in investments_page:
        now = timezone.now()
        total_days = investment.plan.duration_days
        
        days_passed = 0
        if investment.start_date:
            start_date = getattr(investment.start_date, 'date', investment.start_date)()
            days_passed = (now.date() - start_date).days
        
        days_remaining = max(0, total_days - days_passed)
        progress_percentage = min(100, (days_passed / total_days) * 100) if total_days > 0 else 0
        
        investment.days_passed = max(0, days_passed)
        investment.days_remaining = days_remaining
        investment.progress_percentage = progress_percentage
    
    investment_plans = InvestmentPlan.objects.filter(is_active=True)
    users = CustomUser.objects.filter(is_active=True, is_staff=False).order_by('username')
    
    context = {
        'investments': investments_page,
        'investment_plans': investment_plans,
        'users': users,
        'no_plans': not investment_plans.exists(),
        'no_users': not users.exists(),
        **stats,
    }
    
    if not investment_plans.exists() or not users.exists():
        messages.warning(request, 'No active users or investment plans available. Please create some first.')
    
    return render(request, 'admin_panel/investments.html', context)

@admin_required
def manage_investment_plans(request):
    plans = InvestmentPlan.objects.all().order_by('minimum_amount')
    
    for plan in plans:
        plan.total_investments = Investment.objects.filter(plan=plan).count()
        plan.total_invested = Investment.objects.filter(plan=plan).aggregate(Sum('amount'))['amount__sum'] or 0
    
    active_plans = plans.filter(is_active=True).count()
    total_invested = Investment.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    active_investors = Investment.objects.filter(status='active').values('user').distinct().count()
    
    if request.method == 'POST':
        try:
            plan_id = request.POST.get('plan_id')
            if plan_id:
                plan = get_object_or_404(InvestmentPlan, id=plan_id)
            else:
                plan = InvestmentPlan()
            
            plan.name = request.POST.get('name')
            plan.minimum_amount = float(request.POST.get('minimum_amount'))
            plan.maximum_amount = float(request.POST.get('maximum_amount'))
            plan.roi_percentage = float(request.POST.get('roi_percentage'))
            plan.duration_days = int(request.POST.get('duration_days'))
            plan.description = request.POST.get('description')
            plan.is_active = request.POST.get('is_active') == 'true'
            plan.save()
            
            messages.success(request, f'Investment plan "{plan.name}" saved successfully!')
            return redirect('admin_panel:manage_investment_plans')
        except ValueError as e:
            logger.error(f'Error saving investment plan: {str(e)}')
            messages.error(request, f'Invalid input: {str(e)}')
        except Exception as e:
            logger.error(f'Unexpected error saving investment plan: {str(e)}')
            messages.error(request, 'An error occurred while saving the plan.')
    
    context = {
        'plans': plans,
        'active_plans': active_plans,
        'total_invested': total_invested,
        'active_investors': active_investors,
    }
    return render(request, 'admin_panel/investment_plans.html', context)

@admin_required
def settings(request):
    site_settings, created = SiteSettings.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        try:
            site_settings.site_name = request.POST.get('site_name', site_settings.site_name)
            site_settings.contact_email = request.POST.get('site_email', site_settings.contact_email)
            site_settings.btc_wallet = request.POST.get('bitcoin_wallet', site_settings.btc_wallet)
            site_settings.eth_wallet = request.POST.get('ethereum_wallet', site_settings.eth_wallet)
            site_settings.usdt_wallet = request.POST.get('usdt_wallet', site_settings.usdt_wallet)
            
            minimum_deposit = request.POST.get('minimum_deposit')
            if minimum_deposit:
                site_settings.minimum_deposit = float(minimum_deposit)
            
            minimum_withdrawal = request.POST.get('minimum_withdrawal')
            if minimum_withdrawal:
                site_settings.minimum_withdrawal = float(minimum_withdrawal)
            
            site_settings.save()
            messages.success(request, 'Settings updated successfully!')
        except ValueError as e:
            logger.error(f'Error updating settings: {str(e)}')
            messages.error(request, f'Invalid input: {str(e)}')
        except Exception as e:
            logger.error(f'Unexpected error updating settings: {str(e)}')
            messages.error(request, 'An error occurred while updating settings.')
        return redirect('admin_panel:settings')
    
    context = {
        'site_settings': site_settings,
    }
    return render(request, 'admin_panel/settings.html', context)

@admin_required
def reports(request):
    total_users = User.objects.count()
    verified_users = KYCDocument.objects.filter(status='approved').count()
    total_deposits = Deposit.objects.filter(status='approved').aggregate(Sum('amount'))['amount__sum'] or 0
    total_withdrawals = Withdrawal.objects.filter(status='approved').aggregate(Sum('amount'))['amount__sum'] or 0
    active_investments = Investment.objects.filter(status='active').aggregate(Sum('amount'))['amount__sum'] or 0
    
    recent_users = User.objects.order_by('-date_joined')[:10]
    recent_deposits = Deposit.objects.order_by('-created_at')[:10]
    recent_withdrawals = Withdrawal.objects.order_by('-created_at')[:10]
    
    context = {
        'total_users': total_users,
        'verified_users': verified_users,
        'total_deposits': total_deposits,
        'total_withdrawals': total_withdrawals,
        'active_investments': active_investments,
        'recent_users': recent_users,
        'recent_deposits': recent_deposits,
        'recent_withdrawals': recent_withdrawals,
    }
    return render(request, 'admin_panel/reports.html', context)

@admin_required
def toggle_plan_status(request, plan_id):
    plan = get_object_or_404(InvestmentPlan, id=plan_id)
    plan.is_active = not plan.is_active
    plan.save()
    
    status_text = 'activated' if plan.is_active else 'deactivated'
    messages.success(request, f'Investment plan "{plan.name}" has been {status_text}.')
    
    return redirect('admin_panel:manage_investment_plans')

@admin_required
def delete_plan(request, plan_id):
    plan = get_object_or_404(InvestmentPlan, id=plan_id)
    
    active_investments = Investment.objects.filter(plan=plan, status='active').count()
    
    if active_investments > 0:
        messages.error(request, f'Cannot delete plan "{plan.name}" because it has {active_investments} active investments.')
    else:
        plan_name = plan.name
        plan.delete()
        messages.success(request, f'Investment plan "{plan_name}" deleted successfully.')
    
    return redirect('admin_panel:manage_investment_plans')

@login_required
@user_passes_test(is_admin)
def complete_investment(request, investment_id):
    """Mark an investment as completed."""
    try:
        with transaction.atomic():
            investment = get_object_or_404(Investment, id=investment_id)
            if investment.status != 'active':
                return JsonResponse({'success': False, 'error': 'Investment is not active'}, status=400)
            
            investment.status = 'completed'
            investment.completed_at = timezone.now()
            investment.save()
            
            profile = get_or_create_profile(investment.user)
            profile.account_balance += investment.expected_return
            profile.current_investment -= investment.amount
            profile.save()
            
            Transaction.objects.create(
                user=investment.user,
                transaction_type='return',
                amount=investment.expected_return,
                description=f'Return from {investment.plan.name}',
                reference_id=str(investment.id)
            )
            
            return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f'Error completing investment {investment_id}: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@admin_required
def manage_testimonials(request):
    testimonials = Testimonial.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        try:
            testimonial_id = request.POST.get('testimonial_id')
            if testimonial_id:
                testimonial = get_object_or_404(Testimonial, id=testimonial_id)
            else:
                testimonial = Testimonial()
            
            testimonial.name = request.POST.get('name')
            testimonial.position = request.POST.get('position', '')
            testimonial.content = request.POST.get('content')
            testimonial.rating = int(request.POST.get('rating', 5))
            testimonial.is_active = request.POST.get('is_active') == 'on'
            
            if 'avatar' in request.FILES:
                testimonial.avatar = request.FILES['avatar']
            
            testimonial.save()
            messages.success(request, f'Testimonial for "{testimonial.name}" saved successfully!')
        except ValueError as e:
            logger.error(f'Error saving testimonial: {str(e)}')
            messages.error(request, f'Invalid input: {str(e)}')
        except Exception as e:
            logger.error(f'Unexpected error saving testimonial: {str(e)}')
            messages.error(request, 'An error occurred while saving the testimonial.')
        
        return redirect('admin_panel:manage_testimonials')
    
    context = {
        'testimonials': testimonials,
    }
    return render(request, 'admin_panel/testimonials.html', context)

@admin_required
def delete_testimonial(request, testimonial_id):
    testimonial = get_object_or_404(Testimonial, id=testimonial_id)
    testimonial_name = testimonial.name
    testimonial.delete()
    messages.success(request, f'Testimonial for "{testimonial_name}" deleted successfully!')
    return redirect('admin_panel:manage_testimonials')

@admin_required
def toggle_testimonial_status(request, testimonial_id):
    testimonial = get_object_or_404(Testimonial, id=testimonial_id)
    testimonial.is_active = not testimonial.is_active
    testimonial.save()
    status_text = 'activated' if testimonial.is_active else 'deactivated'
    messages.success(request, f'Testimonial for "{testimonial.name}" {status_text}.')
    return redirect('admin_panel:manage_testimonials')

@admin_required
def manage_certifications(request):
    certifications = Certification.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        try:
            certification_id = request.POST.get('certification_id')
            if certification_id:
                certification = get_object_or_404(Certification, id=certification_id)
            else:
                certification = Certification()
            
            certification.name = request.POST.get('name')
            certification.description = request.POST.get('description', '')
            certification.is_active = request.POST.get('is_active') == 'on'
            
            if 'image' in request.FILES:
                certification.image = request.FILES['image']
            
            certification.save()
            messages.success(request, f'Certification "{certification.name}" saved successfully!')
        except ValueError as e:
            logger.error(f'Error saving certification: {str(e)}')
            messages.error(request, f'Invalid input: {str(e)}')
        except Exception as e:
            logger.error(f'Unexpected error saving certification: {str(e)}')
            messages.error(request, 'An error occurred while saving the certification.')
        
        return redirect('admin_panel:manage_certifications')
    
    context = {
        'certifications': certifications,
    }
    return render(request, 'admin_panel/certifications.html', context)

@admin_required
def delete_certification(request, certification_id):
    certification = get_object_or_404(Certification, id=certification_id)
    certification_name = certification.name
    certification.delete()
    messages.success(request, f'Certification "{certification_name}" deleted successfully!')
    return redirect('admin_panel:manage_certifications')

@admin_required
def toggle_certification_status(request, certification_id):
    certification = get_object_or_404(Certification, id=certification_id)
    certification.is_active = not certification.is_active
    certification.save()
    status_text = 'activated' if certification.is_active else 'deactivated'
    messages.success(request, f'Certification "{certification.name}" {status_text}.')
    return redirect('admin_panel:manage_certifications')

@admin_required
def admin_profile(request):
    profile = get_or_create_profile(request.user)
    
    context = {
        'profile': profile,
    }
    return render(request, 'admin_panel/profile.html', context)

@admin_required
@require_http_methods(["POST"])
def update_profile(request):
    try:
        request.user.full_name = request.POST.get('full_name', '')
        request.user.email = request.POST.get('email', '')
        request.user.save()
        
        profile = get_or_create_profile(request.user)
        profile.phone_number = request.POST.get('phone_number', '')
        profile.country = request.POST.get('country', '')
        profile.state = request.POST.get('state', '')
        profile.address = request.POST.get('address', '')
        profile.save()
        
        messages.success(request, 'Profile updated successfully!')
    except Exception as e:
        logger.error(f'Error updating admin profile: {str(e)}')
        messages.error(request, f'Error updating profile: {str(e)}')
    
    return redirect('admin_panel:admin_profile')

@admin_required
@require_http_methods(["POST"])
def admin_change_password(request):
    form = PasswordChangeForm(request.user, request.POST)
    
    if form.is_valid():
        user = form.save()
        update_session_auth_hash(request, user)
        messages.success(request, 'Your password was successfully updated!')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'{field}: {error}')
    
    return redirect('admin_panel:admin_profile')

@login_required
@user_passes_test(is_admin)
def investment_detail(request, investment_id):
    """Get investment details for modal display."""
    try:
        investment = get_object_or_404(Investment, id=investment_id)
        
        now = timezone.now()
        total_days = investment.plan.duration_days
        
        days_passed = 0
        if investment.start_date:
            start_date = getattr(investment.start_date, 'date', investment.start_date)()
            days_passed = (now.date() - start_date).days
        
        days_remaining = max(0, total_days - days_passed)
        progress_percentage = min(100, (days_passed / total_days) * 100) if total_days > 0 else 0
        
        investment.days_passed = max(0, days_passed)
        investment.days_remaining = days_remaining
        investment.progress_percentage = progress_percentage
        investment.profit = investment.expected_return - investment.amount
        
        return render(request, 'admin_panel/investment_detail.html', {
            'investment': investment
        })
    except Exception as e:
        logger.error(f'Error fetching investment details {investment_id}: {str(e)}')
        return HttpResponse('Error loading investment details.', status=500)

@login_required
@user_passes_test(is_admin)
def add_investment(request):
    """Add new investment."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    try:
        with transaction.atomic():
            user_id = request.POST.get('user')
            plan_id = request.POST.get('plan')
            amount = float(request.POST.get('amount'))
            start_date = request.POST.get('start_date')
            
            user = get_object_or_404(CustomUser, id=user_id)
            plan = get_object_or_404(InvestmentPlan, id=plan_id)
            
            if amount < plan.minimum_amount or amount > plan.maximum_amount:
                return JsonResponse({'success': False, 'error': f'Amount must be between ${plan.minimum_amount} and ${plan.maximum_amount}'}, status=400)
            
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d')
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Invalid start date format. Use YYYY-MM-DD.'}, status=400)
            else:
                start_date = timezone.now()
            
            end_date = start_date + timedelta(days=plan.duration_days)
            
            expected_return = amount + (amount * plan.roi_percentage / 100)
            
            investment = Investment.objects.create(
                user=user,
                plan=plan,
                amount=amount,
                expected_return=expected_return,
                start_date=start_date,
                end_date=end_date,
                status='active'
            )
            
            Transaction.objects.create(
                user=user,
                transaction_type='investment',
                amount=-amount,
                description=f'Investment in {plan.name}',
                reference_id=str(investment.id)
            )
            
            profile = get_or_create_profile(user)
            profile.current_investment += amount
            profile.account_balance -= amount
            profile.save()
            
            return JsonResponse({'success': True})
    
    except ValueError as e:
        logger.error(f'ValueError creating investment: {str(e)}')
        return JsonResponse({'success': False, 'error': f'Invalid input: {str(e)}'}, status=400)
    except Exception as e:
        logger.error(f'Unexpected error creating investment: {str(e)}')
        return JsonResponse({'success': False, 'error': 'An error occurred while creating the investment.'}, status=500)
@login_required
@user_passes_test(is_admin)
def edit_investment_form(request, investment_id):
    """Get edit investment form for modal display."""
    try:
        investment = get_object_or_404(Investment, id=investment_id)
        investment_plans = InvestmentPlan.objects.filter(is_active=True)
        
        context = {
            'investment': investment,
            'investment_plans': investment_plans,
            'no_plans': not investment_plans.exists()
        }
        
        if not investment_plans.exists():
            messages.warning(request, 'No active investment plans available.')
        
        return render(request, 'admin_panel/edit_investment.html', context)
    except Exception as e:
        logger.error(f'Error fetching edit investment form {investment_id}: {str(e)}')
        return HttpResponse('Error loading edit form.', status=500)

@login_required
@user_passes_test(is_admin)
def edit_investment(request, investment_id):
    """Edit existing investment."""
    if request.method != 'POST':
        return redirect('admin_panel:manage_investments')
    
    investment = get_object_or_404(Investment, id=investment_id)
    
    try:
        with transaction.atomic():
            old_amount = investment.amount
            old_status = investment.status
            
            plan_id = request.POST.get('plan')
            amount = float(request.POST.get('amount'))
            expected_return = float(request.POST.get('expected_return'))
            status = request.POST.get('status')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            completed_at = request.POST.get('completed_at')
            
            plan = get_object_or_404(InvestmentPlan, id=plan_id)
            
            if amount < plan.minimum_amount or amount > plan.maximum_amount:
                messages.error(request, f'Amount must be between ${plan.minimum_amount} and ${plan.maximum_amount}')
                return redirect('admin_panel:manage_investments')
            
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
                end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
                if completed_at and status == 'completed':
                    completed_at = datetime.strptime(completed_at, '%Y-%m-%dT%H:%M')
                else:
                    completed_at = None
            except ValueError:
                messages.error(request, 'Invalid date format. Use YYYY-MM-DDTHH:MM.')
                return redirect('admin_panel:manage_investments')
            
            valid_transitions = {
                'active': ['completed', 'cancelled'],
                'completed': [],
                'cancelled': []
            }
            if old_status != status and status not in valid_transitions[old_status]:
                messages.error(request, f'Invalid status transition from {old_status} to {status}')
                return redirect('admin_panel:manage_investments')
            
            investment.plan = plan
            investment.amount = amount
            investment.expected_return = expected_return
            investment.status = status
            investment.start_date = start_date
            investment.end_date = end_date
            investment.completed_at = completed_at
            investment.save()
            
            profile = get_or_create_profile(investment.user)
            
            if old_amount != amount:
                difference = amount - old_amount
                profile.current_investment += difference
                profile.save()
            
            if old_status != status:
                if status == 'completed' and old_status == 'active':
                    profile.account_balance += investment.expected_return
                    profile.current_investment -= investment.amount
                    profile.save()
                    
                    Transaction.objects.create(
                        user=investment.user,
                        transaction_type='roi',
                        amount=investment.expected_return,
                        description=f'ROI from {investment.plan.name}',
                        reference_id=str(investment.id)
                    )
                elif status == 'cancelled' and old_status == 'active':
                    profile.account_balance += investment.amount
                    profile.current_investment -= investment.amount
                    profile.save()
                    
                    Transaction.objects.create(
                        user=investment.user,
                        transaction_type='refund',
                        amount=investment.amount,
                        description=f'Refund for cancelled investment in {investment.plan.name}',
                        reference_id=str(investment.id)
                    )
            
            messages.success(request, 'Investment updated successfully')
    
    except ValueError as e:
        logger.error(f'ValueError updating investment {investment_id}: {str(e)}')
        messages.error(request, f'Invalid input: {str(e)}')
    except Exception as e:
        logger.error(f'Unexpected error updating investment {investment_id}: {str(e)}')
        messages.error(request, 'An error occurred while updating the investment.')
    
    return redirect('admin_panel:manage_investments')

@login_required
@user_passes_test(is_admin)
def cancel_investment(request, investment_id):
    """Cancel an investment and refund the amount."""
    try:
        with transaction.atomic():
            investment = get_object_or_404(Investment, id=investment_id)
            if investment.status != 'active':
                return JsonResponse({'success': False, 'error': 'Investment is not active'}, status=400)
            
            investment.status = 'cancelled'
            investment.completed_at = timezone.now()
            investment.save()
            
            profile = get_or_create_profile(investment.user)
            profile.account_balance += investment.amount
            profile.current_investment -= investment.amount
            profile.save()
            
            Transaction.objects.create(
                user=investment.user,
                transaction_type='refund',
                amount=investment.amount,
                description=f'Refunded from {investment.plan.name}',
                reference_id=str(investment.id)
            )
            
            return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f'Error cancelling investment {investment_id}: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
        
@login_required
@user_passes_test(is_admin)
def investment_analytics(request):
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    investments = Investment.objects.filter(start_date__gte=start_date)
    
    chart_data = []
    new_investments_data = []
    completed_investments_data = []
    chart_labels = []
    
    for i in range(days):
        date = start_date + timedelta(days=i)
        date_str = date.strftime('%Y-%m-%d')
        chart_labels.append(date.strftime('%m/%d'))
        
        new_count = investments.filter(start_date__date=date.date()).count()
        completed_count = investments.filter(completed_at__date=date.date()).count()
        
        new_investments_data.append(new_count)
        completed_investments_data.append(completed_count)
    
    status_distribution = [
        Investment.objects.filter(status='active').count(),
        Investment.objects.filter(status='completed').count(),
        Investment.objects.filter(status='cancelled').count(),
    ]
    
    plan_performance = InvestmentPlan.objects.annotate(
        total_invested=Sum('investment__amount'),
        investment_count=Count('investment')
    ).filter(total_invested__isnull=False)
    
    plan_names = [plan.name for plan in plan_performance]
    plan_amounts = [float(plan.total_invested or 0) for plan in plan_performance]
    
    top_investors = CustomUser.objects.annotate(
        total_invested=Sum('investments__amount'),
        active_count=Count('investments', filter=Q(investments__status='active'))
    ).filter(total_invested__isnull=False).order_by('-total_invested')[:10]
    
    recent_investments = Investment.objects.select_related('user', 'plan').order_by('-start_date')[:10]
    recent_activities = []
    
    for inv in recent_investments:
        if inv.status == 'completed':
            recent_activities.append({
                'title': 'Investment Completed',
                'description': f'{inv.user.username} completed investment in {inv.plan.name}',
                'created_at': inv.completed_at or inv.start_date
            })
        else:
            recent_activities.append({
                'title': 'New Investment',
                'description': f'{inv.user.username} invested ${inv.amount} in {inv.plan.name}',
                'created_at': inv.start_date
            })
    
    context = {
        'chart_labels': json.dumps(chart_labels),
        'new_investments_data': json.dumps(new_investments_data),
        'completed_investments_data': json.dumps(completed_investments_data),
        'status_distribution': json.dumps(status_distribution),
        'plan_names': json.dumps(plan_names),
        'plan_amounts': json.dumps(plan_amounts),
        'top_investors': top_investors,
        'recent_activities': recent_activities,
    }
    
    return render(request, 'admin_panel/investment_analytics.html', context)

def export_investments(investments, format_type):
    """Export investments to CSV or Excel."""
    try:
        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="investments.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'ID', 'User', 'Email', 'Plan', 'Amount', 'Expected Return',
                'Status', 'Start Date', 'End Date', 'Completed At'
            ])
            
            for investment in investments:
                writer.writerow([
                    investment.id,
                    investment.user.full_name or investment.user.username,
                    investment.user.email,
                    investment.plan.name,
                    investment.amount,
                    investment.expected_return,
                    investment.status,
                    investment.start_date,
                    investment.end_date,
                    investment.completed_at or ''
                ])
            
            return response
        
        elif format_type == 'excel':
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="investments.xlsx"'
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Investments'
            
            headers = [
                'ID', 'User', 'Email', 'Plan', 'Amount', 'Expected Return',
                'Status', 'Start Date', 'End Date', 'Completed At'
            ]
            
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                worksheet[f'{col_letter}1'] = header
            
            for row_num, investment in enumerate(investments, 2):
                worksheet[f'A{row_num}'] = investment.id
                worksheet[f'B{row_num}'] = investment.user.full_name or investment.user.username
                worksheet[f'C{row_num}'] = investment.user.email
                worksheet[f'D{row_num}'] = investment.plan.name
                worksheet[f'E{row_num}'] = float(investment.amount)
                worksheet[f'F{row_num}'] = float(investment.expected_return)
                worksheet[f'G{row_num}'] = investment.status
                worksheet[f'H{row_num}'] = investment.start_date
                worksheet[f'I{row_num}'] = investment.end_date
                worksheet[f'J{row_num}'] = investment.completed_at or ''
            
            workbook.save(response)
            return response
    except Exception as e:
        logger.error(f'Error exporting investments: {str(e)}')
        return HttpResponse('Error exporting investments.', status=500)